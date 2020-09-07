import openpyxl


def getRowCount(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def getColumnCount(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def readData(file_path, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number, column=column_number).value


def writeData(file_path, sheet_name, row_number, column_number, data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number, column=column_number).value = data
    workbook.save(file_path)


